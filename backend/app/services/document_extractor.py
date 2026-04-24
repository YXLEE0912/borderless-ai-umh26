from __future__ import annotations

import asyncio
import base64
import io
import json
import re

import httpx
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

from app.schemas.documents import DocumentExtractedData, DocumentExtractionResponse
from app.services.zai_client import ZAIClient


def _coerce_positive_float(value: float | None) -> float | None:
    return value if value is not None and value > 0 else None


def normalize_document_data(data: DocumentExtractedData, notes: list[str] | None = None) -> DocumentExtractedData:
    quantity = data.quantity if data.quantity is not None and data.quantity > 0 else None
    weight_kg = data.weight_kg if data.weight_kg is not None and data.weight_kg > 0 else None
    declared_value = data.declared_value if data.declared_value is not None and data.declared_value > 0 else None
    unit_price = data.unit_price if data.unit_price is not None and data.unit_price > 0 else None

    if declared_value is not None and quantity is not None and unit_price is not None:
        expected_declared_value = round(quantity * unit_price, 2)
        if abs(expected_declared_value - declared_value) > 0.01:
            if notes is not None:
                notes.append("Detected price mismatch and recalculated total price from quantity x unit_price.")
            declared_value = expected_declared_value
    elif declared_value is None and quantity is not None and unit_price is not None:
        declared_value = round(quantity * unit_price, 2)
    elif unit_price is None and quantity is not None and declared_value is not None:
        unit_price = round(declared_value / quantity, 2)

    return DocumentExtractedData(
        product_name=data.product_name,
        hs_code=data.hs_code,
        destination_country=data.destination_country,
        destination_address=data.destination_address,
        origin_region=data.origin_region,
        quantity=quantity,
        weight_kg=weight_kg,
        declared_value=declared_value,
        unit_price=unit_price,
        incoterm=data.incoterm,
    )


def _normalize_label_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _looks_like_weight_label(value: str) -> bool:
    normalized = _normalize_label_text(value)
    return normalized in {
        "weight",
        "grossweight",
        "netweight",
        "chargeableweight",
        "billableweight",
        "itemweight",
        "packageweight",
        "shipmentweight",
        "cargoweight",
        "grosswt",
        "netwt",
        "chargeablewt",
        "billablewt",
        "packagewt",
        "weightkg",
        "grossweightkg",
        "netweightkg",
    }


def _looks_like_quantity_label(value: str) -> bool:
    normalized = _normalize_label_text(value)
    return normalized in {
        "quantity",
        "qty",
        "totalqty",
        "itemqty",
        "piece",
        "pieces",
        "unit",
        "units",
        "carton",
        "cartons",
        "box",
        "boxes",
        "pack",
        "packs",
        "package",
        "packages",
        "pcs",
        "count",
        "totalcount",
        "shippingunits",
    }


class DocumentExtractor:
    def __init__(
        self,
        zai_client: ZAIClient,
        has_zai_key: bool,
        google_cloud_api_key: str | None = None,
        gemini_api_keys: list[str] | None = None,
        gemini_model: str = "gemini-2.5-flash",
        openai_api_key: str | None = None,
        openai_model: str = "gpt-4o-mini",
    ):
        self.zai_client = zai_client
        self.has_zai_key = has_zai_key
        self.google_cloud_api_key = google_cloud_api_key
        self.gemini_api_keys = [key for key in (gemini_api_keys or []) if key]
        self.gemini_model = gemini_model
        self.openai_api_key = openai_api_key
        self.openai_model = openai_model

    async def extract(self, file_name: str, mime_type: str | None, content: bytes) -> DocumentExtractionResponse:
        notes: list[str] = []
        text = self._extract_text(file_name=file_name, mime_type=mime_type, content=content)

        is_image = bool(mime_type and mime_type.startswith("image/"))
        # Avoid duplicate OCR round-trips when Z.ai image extraction is available.
        should_use_google_ocr = is_image and self.google_cloud_api_key and not self.has_zai_key
        if should_use_google_ocr:
            try:
                ocr_text = await self._extract_text_with_google_vision(
                    image_bytes=content,
                    image_content_type=mime_type,
                )
                if ocr_text:
                    text = ocr_text
                    notes.append("OCR text extracted from image via Google Vision.")
            except Exception as error:
                notes.append(f"Google Vision OCR fallback used: {error.__class__.__name__}.")

        regex_data = self._extract_with_regex(text)
        regex_data = normalize_document_data(regex_data, notes=notes)

        used_zai = False
        result = regex_data

        if self.has_zai_key:
            try:
                zai_raw = await self.zai_client.extract_document_fields(
                    text_context=text,
                    image_bytes=content if is_image else None,
                    image_content_type=mime_type,
                    file_name=file_name,
                )
                zai_data = parse_document_fields_json(zai_raw)
                result = normalize_document_data(self._merge_data(result, zai_data), notes=notes)
                used_zai = True
                notes.append("Document fields extracted with Z.ai.")
            except Exception as error:
                notes.append(f"Z.ai extraction fallback used: {error.__class__.__name__}.")

                if is_image and self.gemini_api_keys:
                    try:
                        gemini_payload = await self._extract_with_gemini_image(
                            image_bytes=content,
                            image_content_type=mime_type,
                            file_name=file_name,
                        )
                        gemini_ocr_text = str(gemini_payload.get("ocr_text") or "").strip()
                        if gemini_ocr_text:
                            text = gemini_ocr_text
                            notes.append("OCR text extracted from image via Gemini fallback.")

                        gemini_data = parse_document_fields_json(json.dumps(gemini_payload))
                        result = normalize_document_data(self._merge_data(result, gemini_data), notes=notes)
                        notes.append("Document fields extracted with Gemini fallback.")
                    except Exception as gemini_error:
                        notes.append(f"Gemini extraction fallback used: {gemini_error.__class__.__name__}.")

        # Run OpenAI only when Z.ai is unavailable or failed.
        if self.openai_api_key and not used_zai:
            try:
                openai_raw = await self._extract_with_openai(text_context=text, file_name=file_name)
                openai_data = parse_document_fields_json(openai_raw)
                result = normalize_document_data(self._merge_data(result, openai_data), notes=notes)
                notes.append("Document fields extracted with OpenAI.")
            except Exception as error:
                notes.append(f"OpenAI extraction fallback used: {error.__class__.__name__}.")

        preview = text[:500] if text else None
        if not preview:
            notes.append("No text content detected in document.")

        return DocumentExtractionResponse(
            file_name=file_name,
            mime_type=mime_type,
            used_zai=used_zai,
            extracted_text_preview=preview,
            data=result,
            notes=notes,
        )

    def _extract_text(self, file_name: str, mime_type: str | None, content: bytes) -> str:
        lower_name = file_name.lower()
        mime = (mime_type or "").lower()

        if lower_name.endswith((".txt", ".csv", ".json", ".md")) or mime.startswith("text/"):
            return content.decode("utf-8", errors="ignore")

        if lower_name.endswith(".pdf") or mime == "application/pdf":
            reader = PdfReader(io.BytesIO(content))
            return "\n".join((page.extract_text() or "") for page in reader.pages)

        if lower_name.endswith(".docx") or mime in {
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        }:
            document = DocxDocument(io.BytesIO(content))
            return "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())

        if lower_name.endswith(".xlsx") or mime in {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        }:
            workbook = load_workbook(io.BytesIO(content), data_only=True)
            cells: list[str] = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for value in row:
                        if value is None:
                            continue
                        text = str(value).strip()
                        if text:
                            cells.append(text)
            return "\n".join(cells)

        return ""

    async def _extract_text_with_google_vision(self, image_bytes: bytes, image_content_type: str | None) -> str:
        encoded_image = base64.b64encode(image_bytes).decode("ascii")
        payload = {
            "requests": [
                {
                    "image": {"content": encoded_image},
                    "features": [{"type": "TEXT_DETECTION", "maxResults": 1}],
                }
            ]
        }

        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(
                "https://vision.googleapis.com/v1/images:annotate",
                params={"key": self.google_cloud_api_key},
                json=payload,
            )
            response.raise_for_status()
            data = response.json()

        responses = data.get("responses") or []
        if not responses:
            return ""

        first = responses[0] or {}
        text = ""
        full_text = first.get("fullTextAnnotation") or {}
        if isinstance(full_text, dict):
            text = str(full_text.get("text") or "").strip()

        if not text:
            annotations = first.get("textAnnotations") or []
            if annotations and isinstance(annotations, list):
                first_annotation = annotations[0] or {}
                text = str(first_annotation.get("description") or "").strip()

        return text

    async def _extract_with_gemini_image(
        self,
        *,
        image_bytes: bytes,
        image_content_type: str | None,
        file_name: str | None,
    ) -> dict:
        import google.generativeai as genai

        if not self.gemini_api_keys:
            raise RuntimeError("Gemini API keys are not configured")

        instruction = (
            "You are an OCR and field extraction assistant for shipping documents. "
            "Return JSON only with keys: "
            "item, product_name, hs_code, destination_country, destination_address, origin_region, "
            "quantity, weight_kg, unit_price, total_price, incoterm, ocr_text. "
            "origin_region must be west, east, or null. "
            "Use null for missing values. Do not include markdown."
        )

        user_prompt = f"file_name={file_name or 'unknown'}\nExtract OCR text and fields from this document image."

        last_error: Exception | None = None
        mime_candidates = _build_gemini_mime_candidates(image_content_type)

        for api_key in self.gemini_api_keys:
            for mime_type in mime_candidates:
                try:
                    def _call() -> str:
                        genai.configure(api_key=api_key)
                        model = genai.GenerativeModel(self.gemini_model)
                        response = model.generate_content(
                            [
                                instruction,
                                user_prompt,
                                {
                                    "mime_type": mime_type,
                                    "data": image_bytes,
                                },
                            ]
                        )
                        return (response.text or "").strip()

                    raw = await asyncio.to_thread(_call)
                    cleaned = re.sub(r"^```json\s*", "", raw.strip(), flags=re.IGNORECASE)
                    cleaned = re.sub(r"\s*```$", "", cleaned)
                    parsed = json.loads(cleaned)
                    if isinstance(parsed, dict):
                        return parsed
                    raise ValueError("Gemini response is not a JSON object")
                except Exception as error:
                    last_error = error
                    continue

        raise RuntimeError(
            f"Gemini image extraction failed: {last_error.__class__.__name__ if last_error else 'unknown'}"
        )

    async def _extract_with_openai(self, text_context: str | None, file_name: str | None) -> str:
        system_prompt = (
            "You extract shipping and customs fields from business documents. Return valid JSON only with keys: "
            "item, product_name, hs_code, destination_country, destination_address, origin_region, quantity, weight_kg, unit_price, total_price, incoterm. "
            "item and product_name are aliases for the product description. origin_region must be one of: west, east, or null. "
            "If quantity/qty appears, map it to quantity. If a unit price is visible, map it to unit_price. "
            "If only one price is visible, map it to total_price. Use null when missing. No markdown. "
            "Always keep total_price equal to unit_price x quantity when both are present."
        )

        user_prompt = (
            f"file_name={file_name or 'unknown'}\n"
            f"text_context={(text_context or '')[:12000]}\n"
            "Extract the required fields from this document content."
        )

        payload = {
            "model": self.openai_model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "temperature": 0.0,
            "response_format": {"type": "json_object"},
        }

        headers = {"Authorization": f"Bearer {self.openai_api_key}"}
        async with httpx.AsyncClient(base_url="https://api.openai.com/v1", timeout=60.0, headers=headers) as client:
            response = await client.post("/chat/completions", json=payload)
            response.raise_for_status()
            data = response.json()

        return str(data["choices"][0]["message"]["content"] or "").strip()

    def _extract_with_regex(self, text: str) -> DocumentExtractedData:
        raw_text = text or ""
        content = raw_text.replace("\n", " ")

        hs_match = re.search(
            r"(?im)\b(?:hs(?:\s*code)?|tariff(?:\s*code)?)\s*[:=#-]?\s*([0-9]{4}(?:\.[0-9]{2}(?:\.[0-9]{2,4})?)?)\b",
            raw_text,
        )
        if not hs_match:
            hs_match = re.search(r"\b[0-9]{4}\.[0-9]{2}(?:\.[0-9]{2,4})?\b", raw_text)

        quantity_match = re.search(
            r"(?im)^\s*(?:quantity|qty|total\s*qty|total\s*quantity|item\s*qty|item\s*quantity|pcs|pieces|units|unit|cartons|boxes|packages|package|packs|pack|count)\s*[:=]?\s*([0-9][0-9,]*(?:\.\d+)?)\s*(?:pcs|pieces|units|unit|cartons|boxes|packages|package|packs|pack|count)?\s*$",
            raw_text,
        )
        if not quantity_match:
            quantity_match = re.search(
                r"(?:quantity|qty|total\s*qty|total\s*quantity|item\s*qty|item\s*quantity|pcs|pieces|units|unit|cartons|boxes|packages|package|packs|pack|count)\s*[:=]?\s*([0-9][0-9,]*(?:\.\d+)?)",
                content,
                flags=re.IGNORECASE,
            )

        if not quantity_match:
            quantity_match = re.search(
                r"(?im)^\s*([0-9][0-9,]*(?:\.\d+)?)\s*(?:pcs|pieces|units|unit|cartons|boxes|packages|package|packs|pack|count)\s*$",
                raw_text,
            )

        unit_price_match = re.search(
            r"(?im)^\s*(?:unit\s*price|price\s*per\s*(?:item|good|unit)|per\s*unit|item\s*price|unit\s*cost|rate|price)\s*[:=]?\s*([0-9][0-9,]*(?:\.\d+)?)\s*$",
            raw_text,
        )
        if not unit_price_match:
            unit_price_match = re.search(
                r"(?:unit\s*price|price\s*per\s*(?:item|good|unit)|per\s*unit|item\s*price|unit\s*cost|rate|price)\s*[:=]?\s*([0-9][0-9,]*(?:\.\d+)?)",
                content,
                flags=re.IGNORECASE,
            )

        goods_value_match = re.search(
            r"(?im)^\s*(?:goods\s*value|declared\s*value|invoice\s*value|invoice\s*amount|total\s*value|total\s*amount|item\s*value|merchandise\s*value|value|amount|total\s*price)\s*[:=]?\s*([0-9][0-9,]*(?:\.\d+)?)\s*$",
            raw_text,
        )
        if not goods_value_match:
            goods_value_match = re.search(
                r"(?:goods\s*value|declared\s*value|invoice\s*value|invoice\s*amount|total\s*value|total\s*amount|item\s*value|merchandise\s*value|value|amount|total\s*price)\s*[:=]?\s*([0-9][0-9,]*(?:\.\d+)?)",
                content,
                flags=re.IGNORECASE,
            )

        net_weight_match = re.search(
            r"(?im)^\s*(?:net\s*weight|netweight|net\s*wt|net\s*mass|netmass)\s*[:=]?\s*([0-9]+(?:\.\d+)?)\s*(?:kg|kgs|kilograms?)?\b",
            raw_text,
        )
        if not net_weight_match:
            net_weight_match = re.search(
                r"(?:net\s*weight|netweight|net\s*wt|net\s*mass|netmass)\s*[:=]?\s*([0-9]+(?:\.\d+)?)\s*(?:kg|kgs|kilograms?)?",
                content,
                flags=re.IGNORECASE,
            )

        gross_weight_match = re.search(
            r"(?im)^\s*(?:gross\s*weight|grossweight|gross\s*wt|gross\s*mass|grossmass)\s*[:=]?\s*([0-9]+(?:\.\d+)?)\s*(?:kg|kgs|kilograms?)?\b",
            raw_text,
        )
        if not gross_weight_match:
            gross_weight_match = re.search(
                r"(?:gross\s*weight|grossweight|gross\s*wt|gross\s*mass|grossmass)\s*[:=]?\s*([0-9]+(?:\.\d+)?)\s*(?:kg|kgs|kilograms?)?",
                content,
                flags=re.IGNORECASE,
            )

        weight_match = net_weight_match or gross_weight_match
        if not weight_match:
            weight_match = re.search(
                r"(?im)^\s*(?:weight|cargo\s*weight|shipment\s*weight|billable\s*weight|chargeable\s*weight|package\s*weight|item\s*weight|mass|gross\s*wt|net\s*wt|chargeable\s*wt|billable\s*wt|package\s*wt|grossweight|netweight|chargeableweight|billableweight)\s*[:=]?\s*([0-9]+(?:\.\d+)?)\s*(?:kg|kgs|kilograms?)?\b",
                raw_text,
            )
        if not weight_match:
            weight_match = re.search(
                r"(?:weight|cargo\s*weight|shipment\s*weight|billable\s*weight|chargeable\s*weight|package\s*weight|item\s*weight|mass|gross\s*wt|net\s*wt|chargeable\s*wt|billable\s*wt|package\s*wt|grossweight|netweight|chargeableweight|billableweight)\s*[:=]?\s*([0-9]+(?:\.\d+)?)\s*(?:kg|kgs|kilograms?)?",
                content,
                flags=re.IGNORECASE,
            )

        if not weight_match:
            weight_match = re.search(
                r"(?im)^\s*([0-9]+(?:\.\d+)?)\s*(?:kg|kgs|kilograms?)\s*(?:gross\s*weight|net\s*weight|cargo\s*weight|shipment\s*weight|billable\s*weight|chargeable\s*weight|package\s*weight|item\s*weight|weight|gross\s*wt|net\s*wt)\s*$",
                raw_text,
            )

        incoterm_match = re.search(r"\b(EXW|FCA|FOB|CFR|CIF|CPT|CIP|DPU|DAP|DDP)\b", content, flags=re.IGNORECASE)

        destination_match = re.search(
            r"(?im)^\s*(?:destination(?:\s*country)?|country\s*of\s*destination|ship\s*-?\s*to|shipper\s*to|consignee\s*country|delivery\s*country|recipient\s*country|receiver\s*country|to)\s*[:=-]\s*([A-Za-z][A-Za-z ]{1,60})\s*$",
            raw_text,
        )
        if not destination_match:
            destination_match = re.search(
                r"(?:destination(?:\s*country)?|country\s*of\s*destination|ship\s*-?\s*to|shipper\s*to|consignee\s*country|delivery\s*country|recipient\s*country|receiver\s*country)\s*[:=-]\s*([A-Za-z][A-Za-z ]{1,60})",
                content,
                flags=re.IGNORECASE,
            )

        destination_address_match = re.search(
            r"(?im)^\s*(?:complete\s*address|full\s*address|destination\s*address|delivery\s*address|ship\s*-?\s*to\s*address|ship\s*-?\s*to|consignee\s*address|shipping\s*address|delivery\s*location|recipient\s*address|receiver\s*address|consignee\s*details|address\s*line(?:\s*[1-3])?|address)\s*[:=-]\s*([^\n]{3,200})\s*$",
            raw_text,
        )
        if not destination_address_match:
            destination_address_match = re.search(
                r"(?:complete\s*address|full\s*address|destination\s*address|delivery\s*address|ship\s*-?\s*to\s*address|ship\s*-?\s*to|consignee\s*address|shipping\s*address|delivery\s*location|recipient\s*address|receiver\s*address|consignee\s*details|address\s*line(?:\s*[1-3])?)\s*[:=-]\s*([^\n]{3,200})",
                content,
                flags=re.IGNORECASE,
            )

        product_match = re.search(
            r"(?im)^\s*(?:product\s*name|product|item\s*description|item\s*desc|goods\s*description|description|commodity|merchandise|cargo\s*description|goods|item)\s*[:=]\s*([^\n]{2,120})\s*$",
            raw_text,
        )
        if not product_match:
            product_match = re.search(
                r"(?:product\s*name|product|item\s*description|item\s*desc|goods\s*description|description|commodity|merchandise|cargo\s*description|goods|item)\s*[:=]\s*([A-Za-z0-9 .,'\-()]{3,120})",
                content,
                flags=re.IGNORECASE,
            )

        origin_region = self._infer_origin_region(raw_text)

        return DocumentExtractedData(
            product_name=product_match.group(1).strip() if product_match else None,
            hs_code=hs_match.group(0) if hs_match else None,
            destination_country=destination_match.group(1).strip() if destination_match else None,
            destination_address=destination_address_match.group(1).strip() if destination_address_match else None,
            origin_region=origin_region,
            quantity=float(quantity_match.group(1).replace(",", "")) if quantity_match else None,
            weight_kg=float(weight_match.group(1)) if weight_match else None,
            declared_value=float(goods_value_match.group(1).replace(",", "")) if goods_value_match else None,
            unit_price=float(unit_price_match.group(1).replace(",", "")) if unit_price_match else None,
            incoterm=incoterm_match.group(1).upper() if incoterm_match else None,
        )

    def _infer_origin_region(self, content: str):
        lowered = (content or "").lower()

        # Prefer explicit East Malaysia markers first.
        east_markers = (
            "east malaysia",
            "sabah",
            "sarawak",
            "kuching",
            "kota kinabalu",
            "bintulu",
            "miri",
        )
        if any(marker in lowered for marker in east_markers):
            return "east"

        west_markers = (
            "west malaysia",
            "peninsular malaysia",
            "selangor",
            "kuala lumpur",
            "johor",
            "penang",
            "malacca",
            "melaka",
            "perak",
            "kedah",
            "kelantan",
            "terengganu",
            "pahang",
            "negeri sembilan",
            "perlis",
        )
        if any(marker in lowered for marker in west_markers):
            return "west"

        # If document says origin from Malaysia without east-state markers,
        # default to west to match existing rate-card baseline.
        if "origin" in lowered and "malaysia" in lowered:
            return "west"

        return None

    def _merge_data(self, base: DocumentExtractedData, overlay: DocumentExtractedData) -> DocumentExtractedData:
        return DocumentExtractedData(
            product_name=overlay.product_name or base.product_name,
            hs_code=overlay.hs_code or base.hs_code,
            destination_country=overlay.destination_country or base.destination_country,
            destination_address=overlay.destination_address or base.destination_address,
            origin_region=overlay.origin_region or base.origin_region,
            quantity=overlay.quantity if overlay.quantity is not None else base.quantity,
            weight_kg=overlay.weight_kg if overlay.weight_kg is not None else base.weight_kg,
            declared_value=overlay.declared_value if overlay.declared_value is not None else base.declared_value,
            unit_price=overlay.unit_price if overlay.unit_price is not None else base.unit_price,
            incoterm=overlay.incoterm or base.incoterm,
        )


def parse_document_fields_json(content: str) -> DocumentExtractedData:
    text = (content or "").strip()
    if not text:
        raise ValueError("Empty JSON content from Z.ai")

    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)

    data = json.loads(text)
    if not isinstance(data, dict):
        raise ValueError("Expected JSON object for document extraction")

    def _to_float(value: object) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            parsed = float(value)
            return parsed if parsed == parsed else None
        if isinstance(value, str):
            text_value = value.strip()
            if not text_value:
                return None
            cleaned = re.sub(r"[^0-9.+\-]", "", text_value.replace(",", ""))
            if not cleaned:
                return None
            try:
                return float(cleaned)
            except ValueError:
                return None
        return None

    def _num(key: str) -> float | None:
        return _to_float(data.get(key))

    def _num_any(keys: list[str], source: dict | None = None) -> float | None:
        bag = source if isinstance(source, dict) else data
        for key in keys:
            value = _to_float(bag.get(key))
            if value is not None:
                return value
        return None

    def _txt(key: str) -> str | None:
        value = data.get(key)
        if value is None:
            return None
        text_value = str(value).strip()
        return text_value or None

    def _txt_any(keys: list[str]) -> str | None:
        for key in keys:
            text_value = _txt(key)
            if text_value:
                return text_value
        return None

    def _line_item_dicts() -> list[dict]:
        candidates = [
            data.get("line_items"),
            data.get("lineItems"),
            data.get("items"),
            data.get("products"),
            data.get("goods"),
        ]
        rows: list[dict] = []
        for candidate in candidates:
            if isinstance(candidate, list):
                for row in candidate:
                    if isinstance(row, dict):
                        rows.append(row)
        return rows

    line_items = _line_item_dicts()

    line_item_quantity_total = 0.0
    has_line_item_quantity = False
    line_item_weight_total = 0.0
    has_line_item_weight = False
    line_item_declared_value_total = 0.0
    has_line_item_declared_value = False
    line_item_unit_price: float | None = None

    for row in line_items:
        quantity = _num_any([
            "quantity", "qty", "total_qty", "total quantity", "item_qty", "pieces", "units", "pcs", "count"
        ], source=row)
        weight = _num_any([
            "net_weight", "net weight", "weight_kg", "weight kg", "weight", "gross_weight", "gross weight"
        ], source=row)
        unit_price = _num_any([
            "unit_price", "unit price", "price_per_unit", "price per unit", "item_price", "price", "rate"
        ], source=row)
        total_price = _num_any([
            "total_price", "total price", "line_total", "line total", "amount", "value", "declared_value", "invoice_value"
        ], source=row)

        if quantity is not None and quantity > 0:
            has_line_item_quantity = True
            line_item_quantity_total += quantity
        if weight is not None and weight > 0:
            has_line_item_weight = True
            line_item_weight_total += weight
        if total_price is not None and total_price > 0:
            has_line_item_declared_value = True
            line_item_declared_value_total += total_price
        if line_item_unit_price is None and unit_price is not None and unit_price > 0:
            line_item_unit_price = unit_price

        if (
            total_price is None
            and quantity is not None
            and quantity > 0
            and unit_price is not None
            and unit_price > 0
        ):
            has_line_item_declared_value = True
            line_item_declared_value_total += round(quantity * unit_price, 2)

    quantity_value = _num_any([
        "quantity", "qty", "total_qty", "total qty", "total_quantity", "total quantity", "item_qty", "item qty", "pcs", "pieces", "units", "cartons", "boxes"
    ])
    if (quantity_value is None or quantity_value <= 0) and has_line_item_quantity:
        quantity_value = round(line_item_quantity_total, 4)

    weight_value = _num_any([
        "net_weight", "net weight", "gross_weight", "gross weight", "weight_kg", "weight kg", "weight", "billable_weight", "billable weight", "chargeable_weight", "chargeable weight", "cargo_weight", "cargo weight", "net_weight_kg", "gross_weight_kg"
    ])
    if (weight_value is None or weight_value <= 0) and has_line_item_weight:
        weight_value = round(line_item_weight_total, 4)

    declared_value = _num_any([
        "declared_value", "declared value", "total_price", "total price", "price", "invoice_value", "invoice value", "goods_value", "goods value", "total_value", "total value", "item_value", "item value", "amount", "line_total", "line total"
    ])
    if (declared_value is None or declared_value <= 0) and has_line_item_declared_value:
        declared_value = round(line_item_declared_value_total, 2)

    unit_price_value = _num_any([
        "unit_price", "unit price", "price_per_unit", "price per unit", "price_per_item", "price per item", "item_price", "item price", "rate", "cost per unit"
    ])
    if (unit_price_value is None or unit_price_value <= 0) and line_item_unit_price is not None:
        unit_price_value = line_item_unit_price

    if (
        (unit_price_value is None or unit_price_value <= 0)
        and quantity_value is not None
        and quantity_value > 0
        and declared_value is not None
        and declared_value > 0
    ):
        unit_price_value = round(declared_value / quantity_value, 2)

    return DocumentExtractedData(
        product_name=_txt_any(["product_name", "product", "item", "item_description", "item description", "goods_description", "goods description", "description", "commodity", "merchandise", "cargo_description", "cargo description"]),
        hs_code=_txt_any(["hs_code", "hs code", "tariff_code", "tariff code", "customs_code", "customs code"]),
        destination_country=_txt_any(["destination_country", "destination country", "country_of_destination", "country of destination", "destination", "ship_to_country", "ship to country", "consignee_country", "consignee country", "delivery_country", "delivery country"]),
        destination_address=_txt_any(["destination_address", "destination address", "complete_address", "complete address", "full_address", "full address", "delivery_address", "delivery address", "ship_to_address", "ship to address", "consignee_address", "consignee address", "shipping_address", "shipping address", "recipient_address", "recipient address", "receiver_address", "receiver address", "address", "address_line_1", "address line 1", "address_line_2", "address line 2", "address_line_3", "address line 3"]),
        origin_region=_txt_any(["origin_region", "origin region", "origin"]),
        quantity=quantity_value,
        weight_kg=weight_value,
        declared_value=declared_value,
        unit_price=unit_price_value,
        incoterm=_txt_any(["incoterm", "incoterms"]),
    )


def _build_gemini_mime_candidates(image_content_type: str | None) -> list[str]:
    normalized = (image_content_type or "").strip().lower()
    if normalized == "image/jpg":
        normalized = "image/jpeg"

    allowed = {"image/png", "image/jpeg", "image/webp"}
    candidates: list[str] = []

    if normalized in allowed:
        candidates.append(normalized)

    for fallback in ["image/jpeg", "image/png", "image/webp"]:
        if fallback not in candidates:
            candidates.append(fallback)

    return candidates